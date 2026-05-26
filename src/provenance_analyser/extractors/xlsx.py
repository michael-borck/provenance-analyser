"""XLSX metadata extractor — openpyxl workbook.properties + app.xml (same shape as .docx)."""
from __future__ import annotations

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any

_EP_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/extended-properties}"


def extract_xlsx(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    out: dict[str, Any] = {
        "creator_app": None,
        "producer": None,
        "author": None,
        "last_modified_by": None,
        "created": None,
        "modified": None,
        "total_editing_time_minutes": None,
        "revision_count": None,
        "page_count": None,        # interpreted as sheet count for .xlsx
        "word_count": None,
        "paragraph_count": None,
        "title": None,
    }

    # read_only=True is fast; we only need properties, not cell values.
    wb = load_workbook(str(path), read_only=True, data_only=False)
    p = wb.properties
    out["author"] = p.creator or None
    out["last_modified_by"] = p.lastModifiedBy or None
    out["created"] = _iso(p.created)
    out["modified"] = _iso(p.modified)
    out["revision_count"] = int(p.revision) if p.revision else None
    out["title"] = p.title or None

    try:
        out["page_count"] = len(wb.sheetnames)
    except Exception:
        pass

    _merge_app_xml(path, out)

    return out


def _merge_app_xml(path: Path, out: dict[str, Any]) -> None:
    try:
        with zipfile.ZipFile(path) as z:
            with z.open("docProps/app.xml") as f:
                tree = ET.parse(f)
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        return
    root = tree.getroot()

    def _text(tag: str) -> str | None:
        node = root.find(f"{_EP_NS}{tag}")
        return node.text if (node is not None and node.text) else None

    def _int(tag: str) -> int | None:
        t = _text(tag)
        try:
            return int(t) if t is not None else None
        except ValueError:
            return None

    out["creator_app"] = _text("Application") or out["creator_app"]
    out["total_editing_time_minutes"] = _int("TotalTime")


def _iso(dt) -> str | None:
    if dt is None:
        return None
    try:
        return dt.isoformat(timespec="seconds")
    except Exception:
        return str(dt)
